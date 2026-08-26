"""Exportação de relatórios — Excel (openpyxl) e PDF (reportlab).

- XLSX: valores numéricos tipados (não strings), datas corretas, moeda formatada,
  totais opcionais, cabeçalho com título/organização/período/filtros.
- PDF: A4, cabeçalho administrativo (organização, GovFrota, relatório, período,
  filtros, data/hora), rodapé com paginação e usuário gerador, quebra de página
  e repetição de cabeçalho de tabela.
"""

import io
from datetime import datetime
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.timezone import now_local

FONT = "Helvetica"


class RelatorioMeta:
    """Metadados comuns de cabeçalho de relatório."""

    def __init__(
        self,
        *,
        titulo: str,
        organizacao: str,
        periodo: str,
        filtros: list[str] | None = None,
        gerado_em: datetime | None = None,
    ):
        self.titulo = titulo
        self.organizacao = organizacao
        self.periodo = periodo
        self.filtros = [f for f in (filtros or []) if f]
        self.gerado_em = gerado_em or now_local()


# ─────────────────────────────── XLSX ──────────────────────────────────────


def _celula_numero(v):
    if isinstance(v, Decimal):
        return float(v)
    return v


def build_xlsx(
    meta: RelatorioMeta,
    headers: list[str],
    rows: list[list],
    *,
    currency_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
    totals: list | None = None,
) -> bytes:
    """Gera um arquivo .xlsx com formatação básica e tipagem correta."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = meta.titulo[:28] or "Relatório"

    title_font = Font(bold=True, size=14, color="1F2937")
    subtitle_font = Font(size=10, color="6B7280")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = meta.titulo
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws["A2"] = meta.organizacao
    ws["A2"].font = subtitle_font
    ws["A3"] = f"Período: {meta.periodo}"
    ws["A3"].font = subtitle_font
    filtro_texto = "; ".join(meta.filtros) if meta.filtros else "Nenhum"
    ws["A4"] = f"Filtros: {filtro_texto}"
    ws["A4"].font = subtitle_font
    ws["A5"] = f"Gerado em: {meta.gerado_em.strftime('%d/%m/%Y %H:%M')}"
    ws["A5"].font = subtitle_font

    header_row = 7
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = header_row + 1
    for row in rows:
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=_celula_numero(v))
            cell.border = border
            if currency_columns and c - 1 in currency_columns:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif date_columns and c - 1 in date_columns:
                cell.number_format = "dd/mm/yyyy"
        r += 1

    if totals:
        for c, v in enumerate(totals, start=1):
            cell = ws.cell(row=r, column=c, value=_celula_numero(v))
            cell.font = Font(bold=True)
            cell.border = border
            if currency_columns and c - 1 in currency_columns:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Largura automática básica
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────── PDF ───────────────────────────────────────


def _estilos():
    ss = getSampleStyleSheet()
    title = ParagraphStyle(
        "GovFrotaTitle", parent=ss["Title"], fontName=FONT, fontSize=16,
        textColor=colors.HexColor("#1D4ED8"), alignment=TA_LEFT, spaceAfter=2,
    )
    subtitle = ParagraphStyle(
        "GovFrotaSub", parent=ss["Normal"], fontName=FONT, fontSize=9,
        textColor=colors.HexColor("#4B5563"), alignment=TA_LEFT, spaceAfter=1,
    )
    meta = ParagraphStyle(
        "GovFrotaMeta", parent=ss["Normal"], fontName=FONT, fontSize=8,
        textColor=colors.HexColor("#6B7280"), alignment=TA_LEFT,
    )
    table_cell = ParagraphStyle(
        "Cell", parent=ss["Normal"], fontName=FONT, fontSize=8, leading=10,
    )
    table_cell_center = ParagraphStyle(
        "CellC", parent=table_cell, alignment=TA_CENTER,
    )
    table_cell_right = ParagraphStyle(
        "CellR", parent=table_cell, alignment=TA_RIGHT,
    )
    header_cell = ParagraphStyle(
        "HCell", parent=ss["Normal"], fontSize=8.5, leading=11,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    section = ParagraphStyle(
        "Section", parent=ss["Normal"], fontName="Helvetica-Bold", fontSize=11,
        textColor=colors.HexColor("#111827"), spaceBefore=8, spaceAfter=4,
    )
    return title, subtitle, meta, table_cell, table_cell_center, table_cell_right, header_cell, section


def _paginas(canvas, doc, meta, usuario):
    canvas.saveState()
    # Rodapé: paginação + usuário gerador
    canvas.setFont(FONT, 8)
    canvas.setFillColor(colors.HexColor("#6B7280"))
    canvas.drawString(15 * mm, 12 * mm, f"GovFrota · {meta.organizacao}")
    canvas.drawRightString(A4[0] - 15 * mm, 12 * mm, f"Gerado por: {usuario}")
    canvas.drawCentredString(A4[0] / 2, 8 * mm, f"Página {doc.page}")
    canvas.restoreState()


def build_pdf(
    meta: RelatorioMeta,
    headers: list[str],
    rows: list[list],
    *,
    usuario: str = "-",
    currency_columns: set[int] | None = None,
    sections: list[tuple[str, list[str], list[list]]] | None = None,
    totais: list | None = None,
) -> bytes:
    """Gera um PDF A4 profissional com cabeçalho/rodapé e tabelas pagináveis."""
    title, subtitle, meta_style, tcell, tcell_c, tcell_r, hcell, section_style = _estilos()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=18 * mm,
        title=meta.titulo,
        author="GovFrota",
    )

    story = []
    story.append(Paragraph("GovFrota", subtitle))
    story.append(Paragraph(meta.titulo, title))
    story.append(Paragraph(meta.organizacao, subtitle))
    story.append(Paragraph(f"Período: {meta.periodo}", meta_style))
    if meta.filtros:
        story.append(Paragraph(f"Filtros: {'; '.join(meta.filtros)}", meta_style))
    story.append(Paragraph(
        f"Gerado em: {meta.gerado_em.strftime('%d/%m/%Y %H:%M:%S')}",
        meta_style,
    ))
    story.append(Spacer(1, 6))

    def _col_style(i, value):
        if currency_columns and i in currency_columns:
            return tcell_r
        return tcell_c if isinstance(value, (int, float)) else tcell

    def _render_table(hdrs, rws, ttls=None):
        styled_hdr = [Paragraph(f"<b>{h}</b>", hcell) for h in hdrs]
        styled_rows = [
            [Paragraph(str(v), _col_style(i, v)) for i, v in enumerate(row)]
            for row in rws
        ]
        data = [styled_hdr] + styled_rows
        if ttls:
            data.append([Paragraph(f"<b>{v}</b>", _col_style(i, v)) for i, v in enumerate(ttls)])
        col_widths = [max(28 * mm, (A4[0] - 30 * mm) / len(hdrs))] * len(hdrs)
        total_w = sum(col_widths)
        if total_w > A4[0] - 30 * mm:
            scale = (A4[0] - 30 * mm) / total_w
            col_widths = [w * scale for w in col_widths]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        table.setStyle(TableStyle(style))
        return table

    if sections:
        for nome, hdrs, rws in sections:
            story.append(Paragraph(nome, section_style))
            story.append(_render_table(hdrs, rws))
            story.append(Spacer(1, 6))
    else:
        story.append(_render_table(headers, rows, totais))

    doc.build(story, onFirstPage=lambda c, d: _paginas(c, d, meta, usuario),
              onLaterPages=lambda c, d: _paginas(c, d, meta, usuario))
    return buf.getvalue()


# ─────────────────── Consultas comuns de relatório ─────────────────────────


async def get_organizacao_nome(db: AsyncSession, organization_id) -> str:
    from sqlalchemy import select

    from app.models.auth_models import Organization

    org = await db.get(Organization, organization_id)
    return org.name if org else "Organização"
