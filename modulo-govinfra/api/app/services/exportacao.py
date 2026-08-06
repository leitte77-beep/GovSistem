"""Exportação de relatórios em CSV, Excel e PDF (item 45).

A mesma estrutura de dados (`colunas` + `linhas`) alimenta os três formatos, o
que garante que a planilha e o PDF sempre mostrem exatamente o que a tela
mostrou. Excel e PDF dependem de bibliotecas opcionais: se não estiverem
instaladas, a exportação em CSV continua funcionando e a API responde com uma
mensagem clara em vez de erro 500.
"""

import csv
import io
from collections.abc import Sequence
from datetime import date, datetime
from typing import Any

from app.core.errors import AppError


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float):
        return f"{valor:.2f}".replace(".", ",")
    return str(valor)


def para_csv(colunas: Sequence[str], linhas: Sequence[Sequence[Any]]) -> bytes:
    """CSV com separador ponto e vírgula e BOM — abre direto no Excel em pt-BR."""
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    escritor.writerow(colunas)
    for linha in linhas:
        escritor.writerow([_texto(valor) for valor in linha])
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def para_excel(
    colunas: Sequence[str], linhas: Sequence[Sequence[Any]], titulo: str = "Relatório"
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover - ambiente sem a dependência opcional
        raise AppError(
            "A exportação em Excel não está disponível nesta instalação. "
            "Use CSV ou peça ao administrador para instalar a biblioteca openpyxl.",
            503,
            "exportacao_indisponivel",
        )

    planilha = Workbook()
    aba = planilha.active
    aba.title = titulo[:31] or "Relatório"

    aba.append(list(colunas))
    cabecalho_fundo = PatternFill("solid", fgColor="1E3A8A")
    for celula in aba[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = cabecalho_fundo
        celula.alignment = Alignment(vertical="center")

    for linha in linhas:
        aba.append([_valor_excel(v) for v in linha])

    for indice, coluna in enumerate(colunas, start=1):
        largura = max(
            len(str(coluna)),
            *(len(_texto(linha[indice - 1])) for linha in linhas[:200] or [[""] * len(colunas)]),
        )
        aba.column_dimensions[get_column_letter(indice)].width = min(max(largura + 2, 12), 60)

    aba.freeze_panes = "A2"
    aba.auto_filter.ref = aba.dimensions

    saida = io.BytesIO()
    planilha.save(saida)
    return saida.getvalue()


def _valor_excel(valor: Any) -> Any:
    """Preserva número e data como tipo nativo — a planilha precisa somar."""
    if isinstance(valor, (int, float, datetime, date)) and not isinstance(valor, bool):
        return valor
    return _texto(valor)


def para_pdf(
    colunas: Sequence[str],
    linhas: Sequence[Sequence[Any]],
    *,
    titulo: str,
    subtitulo: str | None = None,
    rodape: str | None = None,
    paisagem: bool = True,
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:  # pragma: no cover - ambiente sem a dependência opcional
        raise AppError(
            "A exportação em PDF não está disponível nesta instalação. "
            "Use CSV ou peça ao administrador para instalar a biblioteca reportlab.",
            503,
            "exportacao_indisponivel",
        )

    saida = io.BytesIO()
    tamanho = landscape(A4) if paisagem else A4
    documento = SimpleDocTemplate(
        saida,
        pagesize=tamanho,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=titulo,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloGovInfra", parent=estilos["Title"], fontSize=15, spaceAfter=2, alignment=0
    )
    estilo_sub = ParagraphStyle(
        "SubGovInfra", parent=estilos["Normal"], fontSize=9, textColor=colors.HexColor("#475569")
    )
    estilo_celula = ParagraphStyle(
        "CelulaGovInfra", parent=estilos["Normal"], fontSize=7.5, leading=9.5
    )

    elementos = [Paragraph(titulo, estilo_titulo)]
    if subtitulo:
        elementos.append(Paragraph(subtitulo, estilo_sub))
    elementos.append(
        Paragraph(
            f"Emitido em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", estilo_sub
        )
    )
    elementos.append(Spacer(1, 8))

    dados = [[Paragraph(f"<b>{c}</b>", estilo_celula) for c in colunas]]
    for linha in linhas:
        dados.append([Paragraph(_texto(valor), estilo_celula) for valor in linha])

    if len(dados) == 1:
        elementos.append(
            Paragraph("Nenhum registro encontrado para os filtros informados.", estilos["Normal"])
        )
    else:
        tabela = Table(dados, repeatRows=1, hAlign="LEFT")
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elementos.append(tabela)

    if rodape:
        elementos.append(Spacer(1, 10))
        elementos.append(Paragraph(rodape, estilo_sub))

    documento.build(elementos)
    return saida.getvalue()


TIPOS_MIME = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def gerar(
    formato: str,
    colunas: Sequence[str],
    linhas: Sequence[Sequence[Any]],
    *,
    titulo: str,
    subtitulo: str | None = None,
) -> tuple[bytes, str, str]:
    """Gera o arquivo. Retorna (conteúdo, mime, nome sugerido)."""
    formato = (formato or "csv").lower()
    base = _nome_arquivo(titulo)

    if formato == "csv":
        return para_csv(colunas, linhas), TIPOS_MIME["csv"], f"{base}.csv"
    if formato in {"xlsx", "excel"}:
        return para_excel(colunas, linhas, titulo), TIPOS_MIME["xlsx"], f"{base}.xlsx"
    if formato == "pdf":
        return (
            para_pdf(colunas, linhas, titulo=titulo, subtitulo=subtitulo),
            TIPOS_MIME["pdf"],
            f"{base}.pdf",
        )
    raise AppError(
        f"Formato de exportação não suportado: {formato}. Use csv, xlsx ou pdf.",
        422,
        "formato_invalido",
    )


def _nome_arquivo(titulo: str) -> str:
    import re
    import unicodedata

    normalizado = unicodedata.normalize("NFKD", titulo)
    limpo = "".join(c for c in normalizado if not unicodedata.combining(c))
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", limpo).strip("-").lower()
    return f"govinfra-{slug or 'relatorio'}-{date.today().isoformat()}"
