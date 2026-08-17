"""Engine de execucao de relatorios customizaveis (Fases 3.17-3.18).

Cada "fonte de dados" e uma entrada fixa em FONTE_REGISTRY, resolvida para um
modelo SQLAlchemy real e um conjunto whitelisted de campos->colunas. O cliente
nunca envia SQL: `fonte_dados` so pode referenciar uma `tabela` conhecida, e
`colunas`/`filtros`/`ordenacao`/`agrupamentos` so podem referenciar campos
daquela fonte. O filtro de tenant_id e sempre aplicado pelo engine, nunca pelo
chamador — não ha caminho para um relatorio ler dados de outro tenant.

Suporta:
- Selecao/filtro/ordenacao parametrizados sobre um modelo whitelisted
- Agrupamento com totais e porcentagens (em Python, pos-query)
- Exportacao: PDF (WeasyPrint HTML), CSV, Excel (openpyxl), JSON
"""

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

from sqlalchemy import Date, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql.elements import ColumnElement

from app.models.acao_coletiva import AcaoColetiva
from app.models.acolhimento import Acolhimento
from app.models.attendance import Attendance
from app.models.beneficio import ConcessaoBeneficio
from app.models.encaminhamento import Encaminhamento
from app.models.family import Family
from app.models.ivs import IvsCalculo
from app.models.person import Person
from app.models.unit import Unit

logger = logging.getLogger("govsocial.reports")


class RelatorioConfigInvalidoError(ValueError):
    """Configuracao de relatorio referencia fonte/campo desconhecido ou SQL customizado."""


@dataclass(frozen=True)
class CampoFonte:
    titulo: str
    coluna: ColumnElement


@dataclass(frozen=True)
class FonteConfig:
    label: str
    model: Any
    campos: dict[str, CampoFonte]
    tem_soft_delete: bool = True
    tenant_id_como_texto: bool = False
    joins: tuple = field(default_factory=tuple)


def _construir_registro() -> dict[str, FonteConfig]:
    return {
        "families": FonteConfig(
            label="Famílias", model=Family, tenant_id_como_texto=False,
            joins=((Person, Family.responsavel_id == Person.id),),
            campos={
                "id": CampoFonte("ID", Family.id),
                "codigo_familiar": CampoFonte("Código Familiar", Family.codigo),
                "nis_responsavel": CampoFonte("NIS Responsável", Family.nis_responsavel),
                "nome_responsavel": CampoFonte(
                    "Responsável Familiar",
                    func.coalesce(Person.nome_social, Person.nome_civil),
                ),
                "data_cadastro": CampoFonte("Data Cadastro", Family.created_at),
                "situacao_rua": CampoFonte("Situação de Rua", Family.situacao_rua),
                "bairro": CampoFonte("Bairro", Family.bairro),
                "logradouro": CampoFonte("Logradouro", Family.logradouro),
            },
        ),
        "persons": FonteConfig(
            label="Pessoas", model=Person,
            campos={
                "id": CampoFonte("ID", Person.id),
                "nome": CampoFonte("Nome", Person.nome_civil),
                "nome_social": CampoFonte("Nome Social", Person.nome_social),
                "cpf": CampoFonte("CPF", Person.cpf),
                "nis": CampoFonte("NIS", Person.nis),
                "data_nascimento": CampoFonte("Data Nascimento", Person.data_nascimento),
                "sexo": CampoFonte("Sexo", Person.sexo),
                "escolaridade": CampoFonte("Escolaridade", Person.escolaridade),
                "tipo_deficiencia": CampoFonte("Deficiência", Person.tipo_deficiencia),
            },
        ),
        "attendances": FonteConfig(
            label="Atendimentos", model=Attendance,
            campos={
                "id": CampoFonte("ID", Attendance.id),
                "data_atendimento": CampoFonte("Data", Attendance.data_atendimento),
                "tipo": CampoFonte("Tipo", Attendance.tipo),
                "service_type_code": CampoFonte("Serviço", Attendance.service_type_code),
                "unit_id": CampoFonte("Unidade", Attendance.unit_id),
            },
        ),
        "benefit_concessions": FonteConfig(
            label="Benefícios", model=ConcessaoBeneficio, tem_soft_delete=False,
            campos={
                "id": CampoFonte("ID", ConcessaoBeneficio.id),
                "benefit_type_code": CampoFonte("Tipo Benefício", ConcessaoBeneficio.benefit_type_code),
                "status": CampoFonte("Status", ConcessaoBeneficio.status),
                "data_solicitacao": CampoFonte("Data Solicitação", ConcessaoBeneficio.data_solicitacao),
                "quantidade": CampoFonte("Qtd Autorizada", ConcessaoBeneficio.quantidade),
                "valor_total": CampoFonte("Valor Total", ConcessaoBeneficio.valor_total),
                "unit_id": CampoFonte("Unidade", ConcessaoBeneficio.unit_id),
            },
        ),
        "units": FonteConfig(
            label="Unidades", model=Unit,
            campos={
                "id": CampoFonte("ID", Unit.id),
                "nome": CampoFonte("Nome", Unit.nome),
                "tipo": CampoFonte("Tipo", Unit.tipo),
                "cnpj": CampoFonte("CNPJ", Unit.cnpj),
            },
        ),
        "encaminhamentos": FonteConfig(
            label="Encaminhamentos", model=Encaminhamento,
            campos={
                "id": CampoFonte("ID", Encaminhamento.id),
                "tipo": CampoFonte("Tipo", Encaminhamento.tipo),
                "status": CampoFonte("Status", Encaminhamento.status),
                "data_criacao": CampoFonte("Data", Encaminhamento.created_at),
                "unidade_origem_id": CampoFonte("Origem", Encaminhamento.unit_id),
                "unidade_destino": CampoFonte("Destino", Encaminhamento.unidade_destino_id),
            },
        ),
        "acoes_coletivas": FonteConfig(
            label="Grupos/SCFV", model=AcaoColetiva,
            campos={
                "id": CampoFonte("ID", AcaoColetiva.id),
                "nome": CampoFonte("Nome", AcaoColetiva.nome),
                "tipo": CampoFonte("Tipo", AcaoColetiva.tipo),
                "vagas_total": CampoFonte("Vagas Total", AcaoColetiva.vagas_total),
                "vagas_disponiveis": CampoFonte("Vagas Disponíveis", AcaoColetiva.vagas_disponiveis),
            },
        ),
        "acolhimentos": FonteConfig(
            label="Acolhimentos", model=Acolhimento, tenant_id_como_texto=True,
            campos={
                "id": CampoFonte("ID", Acolhimento.id),
                "tipo": CampoFonte("Tipo", Acolhimento.tipo),
                "status": CampoFonte("Status", Acolhimento.status),
                "data_inicio": CampoFonte("Início", Acolhimento.data_inicio),
                "data_fim": CampoFonte("Fim", Acolhimento.data_fim),
                "publico": CampoFonte("Público", Acolhimento.publico),
            },
        ),
        "ivs_calculos": FonteConfig(
            label="IVS", model=IvsCalculo, tem_soft_delete=False, tenant_id_como_texto=True,
            campos={
                "id": CampoFonte("ID", IvsCalculo.id),
                "family_id": CampoFonte("Família", IvsCalculo.family_id),
                "pontuacao": CampoFonte("Pontuação", IvsCalculo.pontuacao),
                "nivel": CampoFonte("Nível", IvsCalculo.nivel),
                "data_calculo": CampoFonte("Data Cálculo", IvsCalculo.data_calculo),
            },
        ),
    }


FONTE_REGISTRY: dict[str, FonteConfig] = _construir_registro()

# Mantido pelo mesmo formato historico (usado por GET /reports/dictionary e
# pelo construtor de relatorios no frontend), agora derivado do registro real
# em vez de mantido a mao — nao ha como o dicionario divergir do que o engine
# de fato sabe executar.
DICIONARIO_DADOS: dict[str, dict] = {
    chave: {
        "label": fonte.label,
        "campos": {campo: cf.titulo for campo, cf in fonte.campos.items()},
    }
    for chave, fonte in FONTE_REGISTRY.items()
}


def _serializar(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.isoformat()
    if isinstance(v, (UUID, Decimal)): return str(v)
    if isinstance(v, bool): return "Sim" if v else "Não"
    return str(v)


def _resolver_fonte(fonte_dados: dict) -> FonteConfig:
    if not isinstance(fonte_dados, dict):
        raise RelatorioConfigInvalidoError("fonte_dados deve ser um objeto")
    if "sql" in fonte_dados:
        raise RelatorioConfigInvalidoError(
            "SQL customizado não é suportado — selecione uma fonte pelo campo "
            "'tabela' (ver GET /reports/dictionary)"
        )
    tabela = fonte_dados.get("tabela")
    if not tabela or tabela not in FONTE_REGISTRY:
        raise RelatorioConfigInvalidoError(f"fonte de dados desconhecida: {tabela!r}")
    return FONTE_REGISTRY[tabela]


def _validar_campos(fonte: FonteConfig, campos: list[str], contexto: str) -> None:
    desconhecidos = sorted({c for c in campos if c not in fonte.campos})
    if desconhecidos:
        raise RelatorioConfigInvalidoError(
            f"{contexto}: campo(s) inválido(s) para a fonte '{fonte.label}': {desconhecidos}"
        )


def validar_config_relatorio(config: dict) -> FonteConfig:
    """Valida fonte_dados/colunas/filtros/ordenacao/agrupamentos contra o
    registro de fontes conhecidas. Levanta RelatorioConfigInvalidoError se algo
    referenciar SQL customizado ou um campo fora do allowlist da fonte.
    """
    fonte = _resolver_fonte(config.get("fonte_dados") or {})

    colunas = config.get("colunas") or []
    if not colunas:
        raise RelatorioConfigInvalidoError("relatório precisa de ao menos uma coluna em 'colunas'")
    _validar_campos(fonte, [c["campo"] for c in colunas], "colunas")

    for filtro in config.get("filtros") or []:
        _validar_campos(fonte, [filtro["campo"]], "filtros")
        tipo = filtro.get("tipo")
        if tipo not in (None, "texto", "data", "numero", "select"):
            raise RelatorioConfigInvalidoError(f"filtros: tipo inválido: {tipo!r}")

    for ordem in config.get("ordenacao") or []:
        _validar_campos(fonte, [ordem["campo"]], "ordenacao")
        direcao = (ordem.get("direcao") or "asc").lower()
        if direcao not in ("asc", "desc"):
            raise RelatorioConfigInvalidoError(f"ordenacao: direção inválida: {ordem.get('direcao')!r}")

    for agrup in config.get("agrupamentos") or []:
        _validar_campos(fonte, [agrup["campo"]], "agrupamentos")

    return fonte


async def executar_relatorio(
    db: AsyncSession, config: dict, filtros_params: dict, tenant_id: UUID
) -> list[dict]:
    """Executa um relatorio contra o modelo real da fonte selecionada e
    retorna os dados como lista de dicionarios. O filtro por tenant_id e
    aplicado incondicionalmente pelo engine — nao depende do config recebido.
    """
    fonte = validar_config_relatorio(config)
    campos_selecionados = [c["campo"] for c in config["colunas"]]

    stmt = select(*[fonte.campos[c].coluna.label(c) for c in campos_selecionados]).select_from(fonte.model)
    for alvo, on_clause in fonte.joins:
        stmt = stmt.outerjoin(alvo, on_clause)

    tenant_valor = str(tenant_id) if fonte.tenant_id_como_texto else tenant_id
    stmt = stmt.where(fonte.model.tenant_id == tenant_valor)
    if fonte.tem_soft_delete:
        stmt = stmt.where(fonte.model.deleted_at.is_(None))

    for filtro in config.get("filtros") or []:
        coluna = fonte.campos[filtro["campo"]].coluna
        valor = filtros_params.get(f"filtro_{filtro['campo']}")
        if valor is None and filtro.get("obrigatorio"):
            valor = filtro.get("valor_padrao")
        if valor is None or valor == "":
            continue
        tipo = filtro.get("tipo")
        try:
            if tipo == "data":
                stmt = stmt.where(cast(coluna, Date) == date.fromisoformat(str(valor)))
            elif tipo == "texto":
                stmt = stmt.where(coluna.ilike(f"%{valor}%"))
            elif tipo == "numero":
                stmt = stmt.where(coluna == float(valor))
            else:
                stmt = stmt.where(coluna == valor)
        except (TypeError, ValueError) as exc:
            raise RelatorioConfigInvalidoError(
                f"filtro '{filtro['campo']}': valor inválido para tipo {tipo!r}: {valor!r}"
            ) from exc

    for ordem in config.get("ordenacao") or []:
        coluna = fonte.campos[ordem["campo"]].coluna
        direcao = (ordem.get("direcao") or "asc").lower()
        stmt = stmt.order_by(coluna.desc() if direcao == "desc" else coluna.asc())

    logger.debug("Executando relatorio: fonte=%s tenant=%s colunas=%s", config["fonte_dados"].get("tabela"), tenant_id, campos_selecionados)
    result = await db.execute(stmt)
    rows = result.mappings().all()
    return [{k: _serializar(v) for k, v in row.items()} for row in rows]


def agrupar_dados(dados: list[dict], agrupamentos: list[dict]) -> list[dict]:
    """Aplica agrupamentos com totais e porcentagens."""
    if not agrupamentos or not dados:
        return dados

    campo = agrupamentos[0]["campo"]
    mostrar_totais = agrupamentos[0].get("mostrar_totais", False)
    mostrar_porcentagem = agrupamentos[0].get("mostrar_porcentagem", False)
    total = len(dados)

    grupos: dict[str, list[dict]] = {}
    for d in dados:
        chave = str(d.get(campo, ""))
        grupos.setdefault(chave, []).append(d)

    resultado = []
    for chave, itens in sorted(grupos.items()):
        if mostrar_totais:
            resultado.append({campo: chave, "_tipo": "grupo", "_count": len(itens),
                            "_pct": round(len(itens) / total * 100, 1) if total > 0 and mostrar_porcentagem else None})
        resultado.extend(itens)

    if mostrar_totais:
        resultado.append({"_tipo": "total_geral", "_count": total})
    return resultado


def exportar_csv(dados: list[dict], colunas: list[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c["titulo"] for c in colunas])
    for d in dados:
        writer.writerow([d.get(c["campo"], "") for c in colunas])
    return output.getvalue().encode("utf-8-sig")


def exportar_excel(dados: list[dict], colunas: list[dict], titulo: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, c in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=c["titulo"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, d in enumerate(dados, 2):
            for col_idx, c in enumerate(colunas, 1):
                ws.cell(row=row_idx, column=col_idx, value=d.get(c["campo"], ""))
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        return exportar_csv(dados, colunas)


def gerar_html(dados: list[dict], colunas: list[dict], config: dict, titulo: str) -> str:
    layout = config.get("layout", {})
    zebrado = layout.get("zebrado", True)
    orientacao = layout.get("orientacao", "retrato")
    col_css = " ".join(f"{c.get('largura', 'auto')}" for c in colunas)

    linhas = ""
    for i, d in enumerate(dados):
        bg = "#f9fafb" if zebrado and i % 2 == 0 else "white"
        celulas = "".join(f'<td style="padding:6px 8px;text-align:{c.get("alinhamento","left")}">{d.get(c["campo"],"")}</td>' for c in colunas)
        linhas += f'<tr style="background:{bg}">{celulas}</tr>'

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{titulo}</title>
<style>body{{font-family:Arial,sans-serif;font-size:12px;margin:20px}}h1{{font-size:16px;margin-bottom:8px}}
table{{width:100%;border-collapse:collapse}}th{{background:#1a56db;color:white;padding:8px;text-align:left}}
</style></head><body><h1>{titulo}</h1><table><thead><tr>{
''.join(f'<th>{c["titulo"]}</th>' for c in colunas)
}</tr></thead><tbody>{linhas}</tbody></table>
<p style="margin-top:12px;color:#6b7280;font-size:10px">Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} — Total: {len(dados)} registros</p>
</body></html>"""


def exportar_pdf(dados: list[dict], colunas: list[dict], config: dict, titulo: str) -> bytes:
    """Gera PDF via WeasyPrint. Retorna bytes do PDF."""
    html = gerar_html(dados, colunas, config, titulo)
    try:
        from weasyprint import HTML
        return HTML(string=html).write_pdf()
    except ImportError:
        logger.warning("WeasyPrint indisponivel — exportando HTML em vez de PDF")
        return html.encode("utf-8")
